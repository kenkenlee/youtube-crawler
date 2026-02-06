import csv
import io
from typing import List, Optional
from datetime import datetime
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.models.video import Video

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting video data to various formats"""

    def export_to_csv(self, videos: List[Video]) -> str:
        """
        Export videos to CSV format

        Args:
            videos: List of Video objects to export

        Returns:
            CSV string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([
            'Video ID',
            'Title',
            'Channel',
            'Published Date',
            'Duration (seconds)',
            'Views',
            'Likes',
            'Comments',
            'URL',
            'Has Summary',
            'Summary',
            'Keywords'
        ])

        # Write data
        for video in videos:
            writer.writerow([
                video.video_id,
                video.title,
                video.channel.channel_name if video.channel else '',
                video.published_at.strftime('%Y-%m-%d %H:%M:%S') if video.published_at else '',
                video.duration or 0,
                video.view_count or 0,
                video.like_count or 0,
                video.comment_count or 0,
                video.video_url,
                'Yes' if video.summary else 'No',
                video.summary or '',
                ', '.join(video.channel.keywords) if video.channel and video.channel.keywords else ''
            ])

        return output.getvalue()

    def export_to_excel(self, videos: List[Video]) -> Optional[bytes]:
        """
        Export videos to Excel format

        Args:
            videos: List of Video objects to export

        Returns:
            Excel file bytes or None if openpyxl not available
        """
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl not installed, cannot export to Excel")
            return None

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Videos"

        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header
        headers = [
            'Video ID',
            'Title',
            'Channel',
            'Published Date',
            'Duration (seconds)',
            'Views',
            'Likes',
            'Comments',
            'URL',
            'Has Summary',
            'Summary',
            'Keywords'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data
        for row_num, video in enumerate(videos, 2):
            ws.cell(row=row_num, column=1, value=video.video_id)
            ws.cell(row=row_num, column=2, value=video.title)
            ws.cell(row=row_num, column=3, value=video.channel.channel_name if video.channel else '')
            ws.cell(row=row_num, column=4, value=video.published_at.strftime('%Y-%m-%d %H:%M:%S') if video.published_at else '')
            ws.cell(row=row_num, column=5, value=video.duration or 0)
            ws.cell(row=row_num, column=6, value=video.view_count or 0)
            ws.cell(row=row_num, column=7, value=video.like_count or 0)
            ws.cell(row=row_num, column=8, value=video.comment_count or 0)
            ws.cell(row=row_num, column=9, value=video.video_url)
            ws.cell(row=row_num, column=10, value='Yes' if video.summary else 'No')
            ws.cell(row=row_num, column=11, value=video.summary or '')
            ws.cell(row=row_num, column=12, value=', '.join(video.channel.keywords) if video.channel and video.channel.keywords else '')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_summary_report(self, videos: List[Video]) -> str:
        """
        Generate a summary report in text format

        Args:
            videos: List of Video objects

        Returns:
            Text report
        """
        total_videos = len(videos)
        total_views = sum(v.view_count or 0 for v in videos)
        total_likes = sum(v.like_count or 0 for v in videos)
        total_comments = sum(v.comment_count or 0 for v in videos)
        videos_with_summary = sum(1 for v in videos if v.summary)

        # Get channel breakdown
        channel_stats = {}
        for video in videos:
            if video.channel:
                channel_name = video.channel.channel_name
                if channel_name not in channel_stats:
                    channel_stats[channel_name] = {
                        'count': 0,
                        'views': 0,
                        'likes': 0
                    }
                channel_stats[channel_name]['count'] += 1
                channel_stats[channel_name]['views'] += video.view_count or 0
                channel_stats[channel_name]['likes'] += video.like_count or 0

        # Generate report
        report = []
        report.append("=" * 60)
        report.append("VIDEO COLLECTION SUMMARY REPORT")
        report.append("=" * 60)
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("")
        report.append("OVERVIEW")
        report.append("-" * 60)
        report.append(f"Total Videos: {total_videos:,}")
        report.append(f"Total Views: {total_views:,}")
        report.append(f"Total Likes: {total_likes:,}")
        report.append(f"Total Comments: {total_comments:,}")
        report.append(f"Videos with AI Summary: {videos_with_summary} ({videos_with_summary/total_videos*100:.1f}%)" if total_videos > 0 else "Videos with AI Summary: 0")
        report.append("")
        report.append("CHANNEL BREAKDOWN")
        report.append("-" * 60)

        for channel_name, stats in sorted(channel_stats.items(), key=lambda x: x[1]['views'], reverse=True):
            report.append(f"\n{channel_name}")
            report.append(f"  Videos: {stats['count']:,}")
            report.append(f"  Views: {stats['views']:,}")
            report.append(f"  Likes: {stats['likes']:,}")
            report.append(f"  Avg Views: {stats['views']//stats['count']:,}" if stats['count'] > 0 else "  Avg Views: 0")

        report.append("")
        report.append("=" * 60)

        return "\n".join(report)
